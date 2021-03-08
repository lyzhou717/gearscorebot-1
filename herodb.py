from openpyxl import load_workbook

workbook = load_workbook(filename = 'e7db.xlsx')
sheet = workbook.active

stats = {}
units = list()

for row in sheet.iter_rows(min_row=2, max_row=202, min_col=1, max_col=9, values_only=True):
  name = row[0].lower()
  basestats = {
    'atk':row[1],
    'hp':row[2],
    'spd':row[3],
    'def':row[4],
    'cc':row[5],
    'cd':row[6],
    'eff':row[7],
    'er':row[8]
  }
  stats[name] = basestats
  units.append(name)


def get_hero(name):
  stat = stats.get(name)
  return stat


